from __future__ import annotations

import csv
import io
import json
from pathlib import Path

from apps.api.app.domain.commerce_imports import ImportDataType, ImportPreview


MAX_FILE_SIZE = 20 * 1024 * 1024
SAMPLE_ROW_LIMIT = 5

FIELD_DEFINITIONS: dict[ImportDataType, dict[str, tuple[str, ...]]] = {
    "products": {
        "external_id": ("商品id", "商品编号", "宝贝id", "product_id", "item_id", "id", "productId", "productSn"),
        "title": ("商品名称", "商品标题", "宝贝标题", "title", "name", "productName"),
        "sku": ("商家编码", "sku", "sku编码", "productSkuCode", "productSn"),
        "price": ("价格", "售价", "price", "productPrice", "productRealPrice"),
        "inventory_count": ("库存", "可售库存", "stock", "inventory", "realStock"),
    },
    "orders": {
        "external_id": ("订单号", "主订单号", "order_id", "订单编号", "orderSn", "order_sn"),
        "customer_name": ("买家昵称", "客户名称", "收件人", "buyer_name", "memberUsername", "memberNickname", "receiverName"),
        "status": ("订单状态", "交易状态", "status"),
        "total_amount": ("订单金额", "实付金额", "payment", "total_amount", "payAmount", "totalAmount"),
        "paid_at": ("付款时间", "支付时间", "paid_at", "paymentTime", "createTime"),
    },
    "order_items": {
        "external_id": ("子订单号", "订单明细id", "order_item_id", "id"),
        "order_external_id": ("订单号", "主订单号", "order_id", "orderSn"),
        "title": ("商品名称", "商品标题", "title", "productName"),
        "sku": ("商家编码", "sku", "sku编码", "productSkuCode", "productSn"),
        "quantity": ("数量", "购买数量", "quantity", "productQuantity", "productCount"),
        "unit_price": ("单价", "商品单价", "unit_price", "productPrice", "productRealPrice"),
    },
    "customers": {
        "external_id": ("客户id", "买家id", "buyer_id", "customer_id", "memberId", "memberUsername"),
        "name": ("客户名称", "买家昵称", "昵称", "name", "memberNickname", "memberUsername", "receiverName"),
        "tags": ("标签", "客户标签", "tags"),
    },
    "shipments": {
        "external_id": ("物流单id", "发货单号", "shipment_id"),
        "order_external_id": ("订单号", "主订单号", "order_id", "orderSn"),
        "carrier_name": ("物流公司", "快递公司", "carrier", "deliveryCompany"),
        "tracking_number": ("运单号", "物流单号", "tracking_number", "deliverySn"),
        "status": ("物流状态", "发货状态", "status"),
    },
    "after_sales": {
        "external_id": ("售后单号", "退款单号", "after_sale_id", "returnApplyId", "refund_id", "id"),
        "order_external_id": ("订单号", "主订单号", "order_id", "orderSn"),
        "case_type": ("售后类型", "退款类型", "type", "returnType"),
        "status": ("售后状态", "退款状态", "status"),
        "reason": ("售后原因", "退款原因", "reason", "description"),
    },
}

REQUIRED_FIELDS: dict[ImportDataType, tuple[str, ...]] = {
    "products": ("external_id", "title"),
    "orders": ("external_id", "status", "total_amount"),
    "order_items": ("external_id", "order_external_id", "title", "quantity"),
    "customers": ("external_id", "name"),
    "shipments": ("external_id", "order_external_id", "tracking_number"),
    "after_sales": ("external_id", "order_external_id", "status"),
}


class ImportFileError(ValueError):
    pass


def parse_import_preview(file_name: str, content: bytes, data_type: ImportDataType) -> ImportPreview:
    if not content:
        raise ImportFileError("文件为空，请重新选择文件。")
    if len(content) > MAX_FILE_SIZE:
        raise ImportFileError("文件超过 20MB，请拆分后再导入。")

    suffix = Path(file_name).suffix.lower()
    if suffix == ".csv":
        headers, rows, encoding = _parse_csv(content)
    elif suffix == ".xlsx":
        headers, rows, encoding = _parse_xlsx(content)
    elif suffix == ".json":
        headers, rows, encoding = _parse_json(content)
    else:
        raise ImportFileError("只支持 CSV、.xlsx 或 JSON 文件，不接受宏文件和旧版 .xls。")

    if not headers:
        raise ImportFileError("没有读取到表头，请确认第一行是字段名称。")

    mapping = _suggest_mapping(headers, data_type)
    missing = [field for field in REQUIRED_FIELDS[data_type] if field not in mapping]
    warnings = (f"仍需手动匹配必填字段：{', '.join(missing)}",) if missing else ()
    return ImportPreview(
        file_name=file_name,
        encoding=encoding,
        headers=tuple(headers),
        sample_rows=tuple(rows[:SAMPLE_ROW_LIMIT]),
        suggested_mapping=mapping,
        required_fields=REQUIRED_FIELDS[data_type],
        warnings=warnings,
    )


def parse_import_rows(file_name: str, content: bytes) -> list[dict[str, str]]:
    suffix = Path(file_name).suffix.lower()
    if suffix == ".csv":
        return _parse_csv(content)[1]
    if suffix == ".xlsx":
        return _parse_xlsx(content, row_limit=None)[1]
    if suffix == ".json":
        return _parse_json(content, row_limit=None)[1]
    raise ImportFileError("只支持 CSV、.xlsx 或 JSON 文件。")


def _parse_csv(content: bytes) -> tuple[list[str], list[dict[str, str]], str]:
    last_error: UnicodeDecodeError | None = None
    for encoding in ("utf-8-sig", "utf-8", "gb18030"):
        try:
            text = content.decode(encoding)
            reader = csv.DictReader(io.StringIO(text, newline=""))
            headers = [header.strip() for header in (reader.fieldnames or []) if header and header.strip()]
            rows = [
                {str(key).strip(): "" if value is None else str(value).strip() for key, value in row.items() if key}
                for row in reader
            ]
            return headers, rows, encoding
        except UnicodeDecodeError as error:
            last_error = error
    raise ImportFileError("无法识别 CSV 编码，请使用 UTF-8、UTF-8 BOM 或 GB18030。") from last_error


def _parse_xlsx(content: bytes, row_limit: int | None = SAMPLE_ROW_LIMIT) -> tuple[list[str], list[dict[str, str]], str]:
    try:
        from openpyxl import load_workbook
    except ImportError as error:
        raise ImportFileError("后端尚未安装 Excel 解析依赖，请先安装 openpyxl。") from error

    try:
        workbook = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
        sheet = workbook.active
        values = sheet.iter_rows(values_only=True)
        first_row = next(values, ())
        headers = [str(value).strip() if value is not None else "" for value in first_row]
        rows: list[dict[str, str]] = []
        for values_row in values:
            rows.append(
                {
                    header: "" if value is None else str(value).strip()
                    for header, value in zip(headers, values_row, strict=False)
                    if header
                }
            )
            if row_limit is not None and len(rows) >= row_limit:
                break
        return [header for header in headers if header], rows, "xlsx"
    except Exception as error:
        raise ImportFileError("Excel 文件无法读取，请确认文件未损坏且扩展名为 .xlsx。") from error


def _parse_json(content: bytes, row_limit: int | None = SAMPLE_ROW_LIMIT) -> tuple[list[str], list[dict[str, str]], str]:
    try:
        payload = json.loads(content.decode("utf-8-sig"))
    except (UnicodeDecodeError, json.JSONDecodeError) as error:
        raise ImportFileError("JSON 文件无法读取，请使用 UTF-8 编码并确认格式正确。") from error

    if isinstance(payload, dict):
        for key in ("items", "rows", "data"):
            value = payload.get(key)
            if isinstance(value, list):
                payload = value
                break

    if not isinstance(payload, list):
        raise ImportFileError("JSON 顶层必须是数组，或包含 items、rows、data 数组字段。")

    rows: list[dict[str, str]] = []
    headers: list[str] = []
    seen_headers: set[str] = set()
    for item in payload:
        if not isinstance(item, dict):
            raise ImportFileError("JSON 数组中的每一项都必须是对象。")
        row: dict[str, str] = {}
        for key, value in item.items():
            header = str(key).strip()
            if not header:
                continue
            if header not in seen_headers:
                seen_headers.add(header)
                headers.append(header)
            row[header] = _stringify_json_value(value)
        rows.append(row)
        if row_limit is not None and len(rows) >= row_limit:
            break

    return headers, rows, "json"


def _stringify_json_value(value: object) -> str:
    if value is None:
        return ""
    if isinstance(value, (str, int, float, bool)):
        return str(value).strip()
    return json.dumps(value, ensure_ascii=False, separators=(",", ":"))


def _suggest_mapping(headers: list[str], data_type: ImportDataType) -> dict[str, str]:
    normalized_headers = {header.strip().lower(): header for header in headers}
    result: dict[str, str] = {}
    for target_field, aliases in FIELD_DEFINITIONS[data_type].items():
        for alias in aliases:
            matched = normalized_headers.get(alias.lower())
            if matched:
                result[target_field] = matched
                break
    return result
